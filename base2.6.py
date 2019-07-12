import numpy as np
from openpyxl import load_workbook


def nm(N, M, alpha1, alpha2, A_price1, A_price2, B_price1, B_price2):
    A_n1 = np.empty(N)
    A_n1.fill(1 / (2 * N))
    A_n2 = np.empty(M)
    A_n2.fill(1 / (2 * M))
    B_n2 = np.empty(M)
    B_n2.fill(1 / (2 * M))
    flag = np.arange(1)
    while flag[0] < N + M:
        flag[0] = 0
        for i in range(N):
            xi = 0.5 + 0.5 * (alpha1 * (np.sum(A_n2) - np.sum(B_n2)) -
                              A_price1[i] + B_price1[i])
            if xi < i / N:
                xi = i / N
            elif xi > (i + 1) / N:
                xi = (i + 1) / N
            if A_n1[i] == xi - i / N:
                flag[0] += 1
            else:
                A_n1[i] = xi - i / N
        for j in range(M):
            Ayj = alpha2 * np.sum(A_n1) - A_price2[j]
            Byj = 1 + B_price2[j] - alpha2 * (1 - np.sum(A_n1))
            if Ayj < Byj:
                Ayj = (Ayj + Byj) / 2
                Byj = Ayj
            if Ayj < j / M:
                Ayj = j / M
            elif Ayj > (j + 1) / M:
                Ayj = (j + 1) / M
            if Byj < j / M:
                Byj = j / M
            elif Byj > (j + 1) / M:
                Byj = (j + 1) / M
            if A_n2[j] == Ayj - j / M and B_n2[j] == (j + 1) / M - Byj:
                flag[0] += 1
            else:
                A_n2[j] = Ayj - j / M
                B_n2[j] = (j + 1) / M - Byj
    result = np.concatenate((A_n1, A_n2, B_n2, flag))
    return result


def A_profit(N, M, alpha1, alpha2, A_price1, A_price2, B_price1, B_price2):
    ABnm_temp = nm(N, M, alpha1, alpha2, A_price1,
                   A_price2, B_price1, B_price2)
    A_n1_temp = ABnm_temp[0:N]
    A_n2_temp = ABnm_temp[N:N + M]
    profit_temp = A_price1.T @ A_n1_temp + A_price2.T @ A_n2_temp
    return profit_temp


def B_profit(N, M, alpha1, alpha2, A_price1, A_price2, B_price1, B_price2):
    ABnm_temp = nm(N, M, alpha1, alpha2, A_price1,
                   A_price2, B_price1, B_price2)
    B_n1_temp = 1 / N - ABnm_temp[0:N].copy()
    B_n2_temp = ABnm_temp[N + M:N + 2 * M]
    profit_temp = B_price1.T @ B_n1_temp + B_price2.T @ B_n2_temp
    return profit_temp


def A_profit_k(N, M, alpha1, alpha2, A_price1, A_price2, B_price1, B_price2, k, kind):
    # i从零开始
    step = 1 / N
    accuracy = 1e-3
    A_price1_changed = A_price1.copy()
    A_price2_changed = A_price2.copy()
    if kind == 1:
        A_price1_changed[k] = 0
    else:
        A_price2_changed[k] = 0
    ABnm = nm(N, M, alpha1, alpha2, A_price1_changed,
              A_price2_changed, B_price1, B_price2)
    A_n1 = ABnm[0:N]
    A_n2 = ABnm[N:N + M]
    profit0 = A_price1_changed.T @ A_n1 + A_price2_changed.T @ A_n2
    if kind == 1:
        if A_n1[k] <= 1e-7:
            return 0
        else:
            while A_n1[k] > 0:
                A_price1_changed[k] += step
                profit1 = A_profit(N, M, alpha1, alpha2, A_price1_changed,
                                   A_price2_changed, B_price1, B_price2)
                if profit0 >= profit1:
                    if A_price1_changed[k] == step:
                        # 排除初始情况
                        price_left = 0
                        price_right = step
                    else:
                        price_left = A_price1_changed[k] - 2 * step
                        price_right = A_price1_changed[k]
                    delta_price = price_right - price_left
                    while delta_price > accuracy:
                        # 考察三分之二点的利润
                        A_price1_changed[k] = price_right - delta_price / 3
                        profit1_temp = A_profit(N, M, alpha1, alpha2, A_price1_changed,
                                                A_price2_changed, B_price1, B_price2)
                        # 考察三分之一点的利润
                        A_price1_changed[k] = price_left + delta_price / 3
                        profit0_temp = A_profit(N, M, alpha1, alpha2, A_price1_changed,
                                                A_price2_changed, B_price1, B_price2)
                        if profit0_temp >= profit1_temp:
                            price_right -= delta_price / 3
                        else:
                            price_left += delta_price / 3
                        delta_price = 2 * delta_price / 3
                    break
                else:
                    profit0 = profit1
            return A_price1_changed[k]  # 小数点后三位
    else:
        if A_n2[k] <= 1e-7:
            return 0
        else:
            while A_n2[k] > 0:
                A_price2_changed[k] += step
                profit1 = A_profit(N, M, alpha1, alpha2, A_price1_changed,
                                   A_price2_changed, B_price1, B_price2)
                if profit0 >= profit1:
                    if A_price2_changed[k] == step:
                        # 排除初始情况
                        price_left = 0
                        price_right = step
                    else:
                        price_left = A_price2_changed[k] - 2 * step
                        price_right = A_price2_changed[k]
                    delta_price = price_right - price_left
                    while delta_price > accuracy:
                        # 考察三分之二点的利润
                        A_price2_changed[k] = price_right - delta_price / 3
                        profit1_temp = A_profit(N, M, alpha1, alpha2, A_price1_changed,
                                                A_price2_changed, B_price1, B_price2)
                        # 考察三分之一点的利润
                        A_price2_changed[k] = price_left + delta_price / 3
                        profit0_temp = A_profit(N, M, alpha1, alpha2, A_price1_changed,
                                                A_price2_changed, B_price1, B_price2)
                        if profit0_temp >= profit1_temp:
                            price_right -= delta_price / 3
                        else:
                            price_left += delta_price / 3
                        delta_price = 2 * delta_price / 3
                    break
                else:
                    profit0 = profit1
            return A_price2_changed[k]  # 小数点后三位


def B_profit_k(N, M, alpha1, alpha2, A_price1, A_price2, B_price1, B_price2, k, kind):
    # i从零开始
    step = 1 / N
    accuracy = 1e-3
    B_price1_changed = B_price1.copy()
    B_price2_changed = B_price2.copy()
    if kind == 1:
        B_price1_changed[k] = 0
    else:
        B_price2_changed[k] = 0
    ABnm = nm(N, M, alpha1, alpha2, A_price1,
              A_price2, B_price1_changed, B_price2_changed)
    B_n1 = 1 / N - ABnm[0:N].copy()
    B_n2 = ABnm[N + M:N + 2 * M]
    profit0 = B_price1_changed.T @ B_n1 + B_price2_changed.T @ B_n2
    if kind == 1:
        if B_n1[k] <= 1e-7:
            return 0
        else:
            while B_n1[k] > 0:
                B_price1_changed[k] += step
                profit1 = B_profit(N, M, alpha1, alpha2, A_price1,
                                   A_price2, B_price1_changed, B_price2_changed)
                if profit0 >= profit1:
                    if B_price1_changed[k] == step:
                        # 排除初始情况
                        price_left = 0
                        price_right = step
                    else:
                        price_left = B_price1_changed[k] - 2 * step
                        price_right = B_price1_changed[k]
                    delta_price = price_right - price_left
                    while delta_price > accuracy:
                        # 考察三分之二点的利润
                        B_price1_changed[k] = price_right - delta_price / 3
                        profit1_temp = B_profit(N, M, alpha1, alpha2, A_price1,
                                                A_price2, B_price1_changed, B_price2_changed)
                        # 考察三分之一点的利润
                        B_price1_changed[k] = price_left + delta_price / 3
                        profit0_temp = B_profit(N, M, alpha1, alpha2, A_price1,
                                                A_price2, B_price1_changed, B_price2_changed)
                        if profit0_temp >= profit1_temp:
                            price_right -= delta_price / 3
                        else:
                            price_left += delta_price / 3
                        delta_price = 2 * delta_price / 3
                    break
                else:
                    profit0 = profit1
            return B_price1_changed[k]  # 小数点后三位
    else:
        if B_n2[k] <= 1e-7:
            return 0
        else:
            while B_n2[k] > 0:
                B_price2_changed[k] += step
                profit1 = B_profit(N, M, alpha1, alpha2, A_price1,
                                   A_price2, B_price1_changed, B_price2_changed)
                if profit0 >= profit1:
                    if B_price2_changed[k] == step:
                        # 排除初始情况
                        price_left = 0
                        price_right = step
                    else:
                        price_left = B_price2_changed[k] - 2 * step
                        price_right = B_price2_changed[k]
                    delta_price = price_right - price_left
                    while delta_price > accuracy:
                        # 考察三分之二点的利润
                        B_price2_changed[k] = price_right - delta_price / 3
                        profit1_temp = B_profit(N, M, alpha1, alpha2, A_price1,
                                                A_price2, B_price1_changed, B_price2_changed)
                        # 考察三分之一点的利润
                        B_price2_changed[k] = price_left + delta_price / 3
                        profit0_temp = B_profit(N, M, alpha1, alpha2, A_price1,
                                                A_price2, B_price1_changed, B_price2_changed)
                        if profit0_temp >= profit1_temp:
                            price_right -= delta_price / 3
                        else:
                            price_left += delta_price / 3
                        delta_price = 2 * delta_price / 3
                    break
                else:
                    profit0 = profit1
            return B_price2_changed[k]  # 小数点后三位

N = 1
M = 1
book = load_workbook(filename=r"output.xlsx")
output = book.get_sheet_by_name("24")
A_price1_temp = []
A_price2_temp = []
B_price1_temp = []
B_price2_temp = []
for i in range(1, N + 1):
    A_price1_temp.append(output.cell(row=1, column=i).value)
    B_price1_temp.append(output.cell(row=3, column=i).value)
for j in range(1, M + 1):
    A_price2_temp.append(output.cell(row=2, column=j).value)
    B_price2_temp.append(output.cell(row=4, column=j).value)
A_price1 = np.array(A_price1_temp)
A_price2 = np.array(A_price2_temp)
B_price1 = np.array(B_price1_temp)
B_price2 = np.array(B_price2_temp)
alpha1 = 0.01
alpha2 = 2
iteration = 100  # 迭代次数
for flag in range(iteration):
    for i in range(N):
        A_price1[i] = A_profit_k(N, M, alpha1, alpha2, A_price1, A_price2,
                                 B_price1, B_price2, i, 1)
        print('1')
        B_price1[i] = B_profit_k(N, M, alpha1, alpha2, A_price1, A_price2,
                                 B_price1, B_price2, i, 1)
        print('2', i)
    for j in range(M):
        A_price2[j] = A_profit_k(N, M, alpha1, alpha2, A_price1, A_price2,
                                 B_price1, B_price2, j, 2)
        print('3')
        B_price2[j] = B_profit_k(N, M, alpha1, alpha2, A_price1, A_price2,
                                 B_price1, B_price2, j, 2)
        print('4', j)
    ABnm = nm(N, M, alpha1, alpha2, A_price1, A_price2, B_price1, B_price2)
    print(A_price1, A_price2, B_price1, B_price2, round(np.sum(
        ABnm[0:N]), 3), round(np.sum(ABnm[N:N + M]), 3), round(np.sum(ABnm[N + M:N + 2 * M]), 3))
